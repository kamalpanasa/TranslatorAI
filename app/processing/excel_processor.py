import os
from typing import Callable, Coroutine, Any
from openpyxl import load_workbook
from app.utils.logger import get_logger
from app.utils.exceptions import FileProcessingError

logger = get_logger("app.processing.excel_processor")


class ExcelProcessor:
    @staticmethod
    async def translate_excel_file(
        input_path: str,
        output_path: str,
        translate_func: Callable[[str], Coroutine[Any, Any, str]]
    ) -> str:
        """
        Translates an Excel workbook cell-by-cell:
        1. Open workbook.
        2. Iterate worksheets and cells.
        3. Skip cells starting with "=" (formulas) or empty/non-string cells.
        4. Translate value and replace.
        5. Save workbook.
        """
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Excel file not found: {input_path}")

        try:
            logger.info(f"Opening Excel workbook for translation: {input_path}")
            wb = load_workbook(input_path)
            
            for sheet in wb.worksheets:
                logger.info(f"Translating sheet: {sheet.title}")
                for row in sheet.iter_rows():
                    for cell in row:
                        val = cell.value
                        
                        # Only translate string cells that aren't formulas
                        if val and isinstance(val, str):
                            val_clean = val.strip()
                            if val_clean and not val_clean.startswith("="):
                                translated_val = await translate_func(val_clean)
                                cell.value = translated_val

            # Save the modified spreadsheet
            wb.save(output_path)
            wb.close()
            logger.info(f"Translated Excel workbook saved successfully to: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error processing Excel spreadsheet: {str(e)}", exc_info=True)
            raise FileProcessingError(f"Failed to translate Excel spreadsheet: {str(e)}")
